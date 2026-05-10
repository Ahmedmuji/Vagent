import os
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def sanitize_title(title, existing_titles):
    """Sanitizes the sheet title to meet Excel requirements."""
    # Remove invalid characters: / \ ? * [ ] :
    clean_title = re.sub(r'[\\/\?\*\[\]:]', '', title)
    # Trim to 28 to allow for potential unique suffixes
    base_title = clean_title[:28].strip() or "Sheet"
    
    final_title = base_title
    counter = 1
    while final_title.lower() in existing_titles:
        suffix = f"_{counter}"
        final_title = base_title[:31-len(suffix)] + suffix
        counter += 1
    
    existing_titles.add(final_title.lower())
    return final_title

def create_summary_sheet(wb, sheets_data):
    """Creates a beautiful summary sheet of all recommended hardware."""
    ws = wb.create_sheet(title="RECOMMENDED HARDWARE", index=0)
    
    # Header styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=14)
    
    headers = ["Category", "Requirement Description", "Recommended Fortinet Solution", "Recommended Juniper Solution"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Extract hardware
    hardware_found = []
    for sheet in sheets_data:
        sheet_title = sheet.get("title", "")
        headers_list = sheet.get("headers", [])
        if "References" not in headers_list: continue
        
        ref_idx = headers_list.index("References")
        desc_idx = -1
        for i, h in enumerate(headers_list):
            h_upper = str(h).upper()
            if "DESCRIPTION" in h_upper or "REQUIREMENT" in h_upper or "SPECIFICATION" in h_upper:
                desc_idx = i
                break
        
        if desc_idx == -1: desc_idx = 1 # Fallback
        
        for row in sheet.get("rows", []):
            if isinstance(row, dict) and row.get("row_type") == "data":
                data = row.get("data", [])
                if ref_idx < len(data) and data[ref_idx]:
                    ref_text = str(data[ref_idx])
                    desc_text = str(data[desc_idx]) if desc_idx < len(data) else ""
                    
                    # Parse Fortinet and Juniper
                    parts = ref_text.split('|')
                    f_sol = parts[0].replace('Fortinet:', '').strip() if len(parts) > 0 else ""
                    j_sol = parts[1].replace('Juniper:', '').strip() if len(parts) > 1 else ""
                    
                    hardware_found.append([sheet_title, desc_text, f_sol, j_sol])

    # Write data
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for item in hardware_found:
        ws.append(item)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Hyperlink if URL found
            urls = re.findall(r'https?://[^\s|]+', str(cell.value))
            if urls:
                cell.hyperlink = urls[0]
                cell.font = Font(color="0563C1", underline="single")

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60

def create_formatted_excel(data, output_path):
    """Takes JSON data and generates a polished, premium Excel workbook with section highlighting."""
    print(f"Generating premium Excel file: {output_path}...")
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # --- Premium Styles ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Deep Navy
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Light Gray
    section_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Handle if data is wrapped in 'sheets' or just a list
    sheets_data = data.get("sheets", []) if isinstance(data, dict) else (data if isinstance(data, list) else [data])

    used_titles = set()
    for sheet_data in sheets_data:
        if not isinstance(sheet_data, dict): continue
        
        raw_title = sheet_data.get("title", "Data")
        title = sanitize_title(raw_title, used_titles)
        headers = list(sheet_data.get("headers", []))
        for required_header in ("References", "Admin_Guide_Reference"):
            if required_header not in headers:
                headers.append(required_header)
        rows = sheet_data.get("rows", [])
        if not rows: continue

        ws = wb.create_sheet(title=title)

        # 1. Write Headers
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 2. Write Data
        for row_obj in rows:
            # Handle both list (old) and dict (new) row formats
            row_type = "data"
            row_data = []
            
            if isinstance(row_obj, dict):
                row_type = row_obj.get("row_type", "data")
                row_data = list(row_obj.get("data", []))
            else:
                row_data = list(row_obj)
            while len(row_data) < len(headers):
                row_data.append("")
            if len(row_data) > len(headers):
                row_data = row_data[:len(headers)]

            # Clean row data: Split inline bullets
            clean_row = []
            for cell_value in row_data:
                if isinstance(cell_value, str):
                    processed = re.sub(r'(?<=\S)\s+([a-zA-Z0-9][\.\)])\s+', r'\n\1 ', cell_value)
                    clean_row.append(processed)
                else:
                    clean_row.append(cell_value)
            
            ws.append(clean_row)
            last_row = ws.max_row
            
            # Apply formatting based on row type
            is_section = (row_type == "section")
            for i, cell in enumerate(ws[last_row]):
                cell.border = border
                cell.alignment = alignment
                if is_section:
                    cell.fill = section_fill
                    cell.font = section_font
                
                # Special handling for References column
                header_val = str(headers[i]).strip() if i < len(headers) else ""
                if header_val == "References" and cell.value and not is_section:
                    # Apply light blue fill to highlight references
                    cell.fill = PatternFill(start_color="E1EBF1", end_color="E1EBF1", fill_type="solid")
                    
                    # Try to extract and apply hyperlinks if present
                    text = str(cell.value)
                    urls = re.findall(r'https?://[^\s|]+', text)
                    if urls:
                        # For simplicity, we link to the first URL if multiple exist, 
                        # but we keep the full text visible.
                        cell.hyperlink = urls[0]
                        cell.font = Font(color="0563C1", underline="single")

        # 3. Dynamic Column Adjustment
        is_narrative = "GENERAL CONTENT" in raw_title.upper()
        for col in ws.columns:
            column_letter = col[0].column_letter
            if is_narrative:
                ws.column_dimensions[column_letter].width = 100
            else:
                # Find max length in column
                max_length = 0
                for cell in col:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        line_max = max([len(l) for l in lines])
                        if line_max > max_length: max_length = line_max
                
                header_val = str(headers[col[0].column - 1]).strip() if (col[0].column - 1) < len(headers) else ""
                if header_val in ("References", "Admin_Guide_Reference"):
                    ws.column_dimensions[column_letter].width = 80 # Much wider for references
                elif header_val == "Admin_Guide_Reference_Tag":
                    ws.column_dimensions[column_letter].width = 20
                    ws.column_dimensions[column_letter].hidden = True
                else:
                    adjusted_width = max(15, min(max_length + 2, 65))
                    ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    return True

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Convert JSON extraction to formatted Excel.")
    parser.add_argument("--input", required=True, help="Path to the input JSON file")
    parser.add_argument("--output", help="Optional output path. Defaults to 'data/Extracted Excel Results'")
    args = parser.parse_args()

    # Determine output path
    if args.output:
        final_output = args.output
    else:
        # Default to 'data/Extracted Excel Results' with the same name as input
        output_dir = os.path.join("data", "Extracted Excel Results")
        os.makedirs(output_dir, exist_ok=True)
        filename = os.path.splitext(os.path.basename(args.input))[0] + ".xlsx"
        final_output = os.path.join(output_dir, filename)

    try:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
        create_formatted_excel(data, final_output)
        print(f"SUCCESS: Result saved to {final_output}")
    except Exception as e:
        print(f"Error: {e}")
