"""
admin_guide_enricher.py
=======================
Production-grade module that enriches an already-processed RFP Excel workbook
with authoritative Fortinet Administrator Guide citations.

It reads ``Admin_Guide_Reference_Tag`` values produced by the Gemini extraction
prompt, matches them against a flat TOC index of the FortiOS Administration
Guide using a hybrid semantic + fuzzy + keyword + domain scoring pipeline,
and writes precise citations back into the workbook.

Usage:
    python scripts/admin_guide_enricher.py \\
        --input  "data/Extracted Excel Results/MyRFP.xlsx" \\
        --toc    "output/toc_flat_index.json" \\
        [--output "data/Extracted Excel Results/MyRFP_enriched.xlsx"]
"""

import json
import logging
import os
import re
import time
import html
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rapidfuzz import fuzz

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Optional heavy imports – graceful degradation
# ---------------------------------------------------------------------------
try:
    from sentence_transformers import SentenceTransformer
except ImportError:
    SentenceTransformer = None
    logger.warning("sentence-transformers not installed; semantic matching disabled.")

try:
    from sklearn.metrics.pairwise import cosine_similarity
except ImportError:
    cosine_similarity = None

# Global cache for the machine learning model to drastically speed up processing
GLOBAL_EMBEDDING_MODEL = None

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
TAG_COLUMN_NAME = "Admin_Guide_Reference_Tag"

OUTPUT_COLUMNS = [
    "Admin_Guide_Reference"
]

# Scoring weights
W_SEMANTIC = 0.55
W_FUZZY    = 0.25
W_KEYWORD  = 0.15
W_DOMAIN   = 0.05

# Confidence thresholds
THRESHOLD_AUTO   = 0.60
THRESHOLD_HIGH   = 0.80
THRESHOLD_MEDIUM = 0.55

# Styling
CITATION_FILL  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
REVIEW_FILL    = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
NOT_REQ_FILL   = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
HEADER_FILL    = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT    = Font(color="FFFFFF", bold=True)
BORDER_THIN    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _tokenize(text: str) -> set:
    """Lowercase tokenization."""
    return set(re.findall(r"[a-z0-9]+(?:[-'][a-z0-9]+)*", text.lower()))


class FortinetAdminGuideReferenceEnricher:
    """
    Enriches an RFP Excel workbook with Fortinet Admin Guide citations.

    Workflow:
        1. ``load_workbook()``       – load Excel preserving all formatting
        2. ``load_toc_index()``      – load flat TOC index
        3. ``build_embedding_index()`` – encode TOC titles + build similarity index
        4. ``enrich_workbook()``     – iterate sheets & rows, resolve tags
        5. ``save_output()``         – persist enriched workbook
    """

    def __init__(
        self,
        workbook_path: str,
        toc_index_path: str,
        model_name: str = "all-MiniLM-L6-v2",
        embedding_cache_path: Optional[str] = None,
        admin_guide_pdf_path: Optional[str] = None,
    ):
        self.workbook_path = workbook_path
        self.toc_index_path = toc_index_path
        self.model_name = model_name
        self.embedding_cache_path = embedding_cache_path or os.path.join(
            os.path.dirname(toc_index_path), "toc_embeddings.npz"
        )
        self.admin_guide_pdf_path = admin_guide_pdf_path

        self.wb = None
        self.toc_index: List[Dict[str, Any]] = []
        self.toc_titles: List[str] = []
        self.toc_embeddings: Optional[np.ndarray] = None
        self.model = None

        # Query-level result cache (deduplicates identical lookup_query values)
        self._query_cache: Dict[str, Dict[str, Any]] = {}

        # Statistics
        self.stats = {
            "total_rows": 0,
            "matched": 0,
            "manual_review": 0,
            "not_required": 0,
            "errors": 0,
        }

    # ------------------------------------------------------------------
    # 1. Load workbook
    # ------------------------------------------------------------------
    def load_workbook(self) -> None:
        """Load the Excel workbook preserving all existing data and formatting."""
        logger.info("Loading workbook: %s", self.workbook_path)
        self.wb = load_workbook(self.workbook_path)
        logger.info("Loaded %d sheets: %s", len(self.wb.sheetnames), self.wb.sheetnames)

    # ------------------------------------------------------------------
    # 2. Load TOC index
    # ------------------------------------------------------------------
    def load_toc_index(self) -> None:
        """Load the flat TOC index JSON."""
        logger.info("Loading TOC index: %s", self.toc_index_path)
        with open(self.toc_index_path, "r", encoding="utf-8") as fh:
            self.toc_index = json.load(fh)
        for entry in self.toc_index:
            if isinstance(entry, dict):
                self._normalize_index_entry(entry)
        self.toc_titles = [
            e.get("embedding_text") or e.get("contextual_title") or e.get("section_path") or e.get("path") or e.get("title", "")
            for e in self.toc_index
        ]
        logger.info("Loaded %d TOC entries", len(self.toc_index))

    # ------------------------------------------------------------------
    # 3. Build embedding index
    # ------------------------------------------------------------------
    def build_embedding_index(self) -> None:
        """
        Encode all TOC contextual titles with a sentence-transformer model.
        Caches embeddings to disk for subsequent runs.
        """
        if SentenceTransformer is None:
            logger.warning("sentence-transformers unavailable; skipping embedding index.")
            return

        # Try loading from cache
        if os.path.exists(self.embedding_cache_path):
            logger.info("Loading cached embeddings from %s", self.embedding_cache_path)
            data = np.load(self.embedding_cache_path)
            cached = data["embeddings"]
            if cached.shape[0] == len(self.toc_titles):
                self.toc_embeddings = cached
                logger.info("Loaded cached embeddings (%d × %d)", *cached.shape)
            else:
                logger.warning(
                    "Cache size mismatch (%d vs %d); rebuilding.",
                    cached.shape[0], len(self.toc_titles),
                )

        global GLOBAL_EMBEDDING_MODEL

        if self.toc_embeddings is None:
            logger.info("Encoding %d TOC titles with %s …", len(self.toc_titles), self.model_name)
            if GLOBAL_EMBEDDING_MODEL is None:
                GLOBAL_EMBEDDING_MODEL = SentenceTransformer(self.model_name)
            self.model = GLOBAL_EMBEDDING_MODEL
            self.toc_embeddings = self.model.encode(
                self.toc_titles, show_progress_bar=True, batch_size=128
            ).astype("float32")
            # Persist
            os.makedirs(os.path.dirname(self.embedding_cache_path) or ".", exist_ok=True)
            np.savez_compressed(self.embedding_cache_path, embeddings=self.toc_embeddings)
            logger.info("Embeddings cached to %s", self.embedding_cache_path)

        # Keep model loaded for query encoding
        if self.model is None:
            if GLOBAL_EMBEDDING_MODEL is None:
                logger.info("Loading global SentenceTransformer model into memory...")
                GLOBAL_EMBEDDING_MODEL = SentenceTransformer(self.model_name)
            self.model = GLOBAL_EMBEDDING_MODEL

    # ------------------------------------------------------------------
    # 4. Parse reference tag
    # ------------------------------------------------------------------
    @staticmethod
    def parse_reference_tag(cell_value: Any) -> Optional[Dict[str, Any]]:
        """
        Parse the Admin_Guide_Reference_Tag cell value.

        Returns:
            Parsed dict if the tag contains a valid JSON with
            ``reference_needed=true``, else ``None``.
        """
        if cell_value is None:
            return None

        raw = str(cell_value).strip()
        if not raw or raw in ('""', "''", "None"):
            return None

        # Try direct JSON parse
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict) and parsed.get("reference_needed"):
                return parsed
            return None
        except json.JSONDecodeError:
            pass

        # Try extracting JSON from surrounding text
        match = re.search(r"\{[^{}]+\}", raw)
        if match:
            try:
                parsed = json.loads(match.group())
                if isinstance(parsed, dict) and parsed.get("reference_needed"):
                    return parsed
            except json.JSONDecodeError:
                pass

        return None

    # ------------------------------------------------------------------
    # 5. Scoring helpers
    # ------------------------------------------------------------------
    def _semantic_score(self, query: str, candidate_idx: int) -> float:
        """Cosine similarity between query embedding and candidate embedding."""
        if self.model is None or self.toc_embeddings is None:
            return 0.0
        q_emb = self.model.encode([query]).astype("float32")
        c_emb = self.toc_embeddings[candidate_idx].reshape(1, -1)
        sim = cosine_similarity(q_emb, c_emb)[0][0]
        # Normalize from [-1,1] to [0,1]
        return float(max(0.0, (sim + 1.0) / 2.0))

    @staticmethod
    def _fuzzy_score(query: str, candidate_title: str) -> float:
        """RapidFuzz token-sort ratio normalized to [0, 1]."""
        return fuzz.token_sort_ratio(query.lower(), candidate_title.lower()) / 100.0

    @staticmethod
    def _keyword_score(query_tokens: set, candidate_tokens: set) -> float:
        """Overlap relative to the candidate tokens length (how many of the candidate's keywords are in the query)."""
        if not query_tokens or not candidate_tokens:
            return 0.0
        intersection = query_tokens & candidate_tokens
        return len(intersection) / len(candidate_tokens)

    @staticmethod
    def _domain_boost(tag_domains: List[str], candidate_domains: List[str]) -> float:
        """Overlap between Gemini-assigned domains and TOC-inferred domains. Neutral baseline if tag_domains exist but don't match."""
        if not tag_domains or not candidate_domains:
            return 0.0
        
        # Define some fuzzy domain mappings based on what Gemini often hallucinates
        tag_set = {d.lower() for d in tag_domains}
        if "network security" in tag_set or "next generation firewall" in tag_set:
            tag_set.update({"networking", "routing", "firewall policy", "ips", "antivirus", "web filter"})
        if "remote access" in tag_set:
            tag_set.update({"ipsec vpn", "agentless vpn", "ssl vpn"})
            
        cand_set = {d.lower() for d in candidate_domains}
        intersection = tag_set & cand_set
        if not intersection:
            return 0.2  # Slight baseline so it doesn't penalize completely
        return len(intersection) / max(len(cand_set), 1)

    def score_candidate(
        self,
        query: str,
        query_tokens: set,
        tag_domains: List[str],
        candidate_idx: int,
    ) -> Tuple[float, Dict[str, float]]:
        """
        Compute the hybrid composite score for a single TOC candidate.

        Returns:
            (composite_score, breakdown_dict)
        """
        entry = self.toc_index[candidate_idx]
        c_title = entry.get("contextual_title") or entry.get("section_path") or entry.get("title", "")
        c_tokens = set(entry.get("keywords", []))
        c_domains = entry.get("domains", [])

        sem  = self._semantic_score(query, candidate_idx)
        fuz  = self._fuzzy_score(query, c_title)
        kw   = self._keyword_score(query_tokens, c_tokens)
        dom  = self._domain_boost(tag_domains, c_domains)
        title_tokens = _tokenize(entry.get("title", ""))
        if title_tokens and title_tokens.issubset(query_tokens):
            fuz = max(fuz, 0.85)

        weights = {
            "semantic": W_SEMANTIC if self.model is not None and self.toc_embeddings is not None else 0.0,
            "fuzzy": W_FUZZY,
            "keyword": W_KEYWORD,
            "domain": W_DOMAIN,
        }
        weight_total = sum(weights.values()) or 1.0
        composite = (
            (weights["semantic"] * sem)
            + (weights["fuzzy"] * fuz)
            + (weights["keyword"] * kw)
            + (weights["domain"] * dom)
        ) / weight_total

        breakdown = {"semantic": sem, "fuzzy": fuz, "keyword": kw, "domain": dom}
        return composite, breakdown

    # ------------------------------------------------------------------
    # 6. Find best match
    # ------------------------------------------------------------------
    def find_best_match(
        self, tag_data: Dict[str, Any], requirement_text: str = ""
    ) -> Dict[str, Any]:
        """
        Run the full hybrid retrieval pipeline for one tag.

        Returns dict with keys: title, path, printed_page, pdf_page,
        score, method, confidence, status, rationale
        """
        feature_candidates = tag_data.get("fortinet_feature_candidates") or []
        lookup_query = " ".join(str(item) for item in feature_candidates) if feature_candidates else tag_data.get("lookup_query", "")
        fortinet_domains = tag_data.get("fortinet_domains", [])
        priority = tag_data.get("priority", "medium")

        # Check cache
        cache_key = lookup_query.strip().lower()
        if cache_key in self._query_cache:
            logger.debug("Cache hit for query: %s", lookup_query)
            return self._query_cache[cache_key]

        # Combine query with requirement text for richer context
        combined_query = lookup_query
        if requirement_text and not feature_candidates:
            combined_query = f"{lookup_query} {requirement_text[:200]}"

        query_tokens = _tokenize(combined_query)

        # Pre-filter: get top-K candidates via semantic similarity
        top_k = 15
        candidate_indices = list(range(len(self.toc_index)))

        if self.model is not None and self.toc_embeddings is not None:
            q_emb = self.model.encode([combined_query]).astype("float32")
            sims = cosine_similarity(q_emb, self.toc_embeddings)[0]
            top_indices = np.argsort(sims)[::-1][:top_k]
            candidate_indices = top_indices.tolist()

        # Score each candidate
        best_score = -1.0
        best_idx = -1
        best_breakdown = {}

        for idx in candidate_indices:
            score, breakdown = self.score_candidate(
                combined_query, query_tokens, fortinet_domains, idx
            )
            if score > best_score:
                best_score = score
                best_idx = idx
                best_breakdown = breakdown

        if best_idx < 0:
            result = self._empty_result("No candidates found")
            self._query_cache[cache_key] = result
            return result

        entry = self.toc_index[best_idx]
        pdf_page = entry.get("pdf_page") or entry.get("page", "")
        pdf_anchor = entry.get("pdf_anchor") or entry.get("anchor") or self._page_anchor(pdf_page)
        pdf_uri = self._entry_pdf_uri(entry, pdf_page)

        # Determine method
        methods_active = []
        if best_breakdown.get("semantic", 0) > 0.3:
            methods_active.append("semantic")
        if best_breakdown.get("fuzzy", 0) > 0.3:
            methods_active.append("fuzzy")
        if best_breakdown.get("keyword", 0) > 0.1:
            methods_active.append("keyword")

        if "semantic" in methods_active and "fuzzy" in methods_active:
            method = "hybrid_semantic_fuzzy"
        elif "semantic" in methods_active:
            method = "semantic_only"
        elif "fuzzy" in methods_active:
            method = "fuzzy_only"
        else:
            method = "keyword_only"

        # Confidence
        if best_score >= THRESHOLD_HIGH:
            confidence = "High"
        elif best_score >= THRESHOLD_MEDIUM:
            confidence = "Medium"
        else:
            confidence = "Low"

        # Status
        status = "Matched" if best_score >= THRESHOLD_AUTO else "Manual Review Required"

        # Rationale
        rationale = (
            f"Query: '{lookup_query}' | "
            f"Semantic: {best_breakdown.get('semantic', 0):.3f}, "
            f"Fuzzy: {best_breakdown.get('fuzzy', 0):.3f}, "
            f"Keyword: {best_breakdown.get('keyword', 0):.3f}, "
            f"Domain: {best_breakdown.get('domain', 0):.3f}"
        )

        result = {
            "title": entry["title"],
            "path": entry.get("breadcrumb_path") or entry.get("section_path", ""),
            "printed_page": entry.get("printed_page") or entry.get("page", ""),
            "page": entry.get("page", pdf_page),
            "pdf_page": pdf_page,
            "anchor": pdf_anchor,
            "pdf_anchor": pdf_anchor,
            "pdf_uri": pdf_uri,
            "named_destination": entry.get("named_destination", ""),
            "section_path": entry.get("section_path") or entry.get("breadcrumb_path", ""),
            "score": round(best_score, 4),
            "method": method,
            "confidence": confidence,
            "status": status,
            "rationale": rationale,
        }

        self._query_cache[cache_key] = result
        return result

    @staticmethod
    def _empty_result(reason: str) -> Dict[str, Any]:
        return {
            "title": "",
            "path": "",
            "printed_page": "",
            "page": "",
            "pdf_page": "",
            "anchor": "",
            "pdf_anchor": "",
            "pdf_uri": "",
            "named_destination": "",
            "section_path": "",
            "score": 0.0,
            "method": "",
            "confidence": "Low",
            "status": "Manual Review Required",
            "rationale": reason,
        }

    # ------------------------------------------------------------------
    # 7. Generate citation
    # ------------------------------------------------------------------
    @staticmethod
    def generate_citation(match: Dict[str, Any]) -> str:
        """
        Format a citation string.

        Example:
            Fortinet Administrator Guide, "High Availability", p. 845 (PDF p. 867)
        """
        title = match.get("title", "")
        printed = match.get("printed_page", "")
        pdf = match.get("pdf_page", "")
        if not title:
            return ""
        return f"FortiOS Admin Guide - {title} (Page {pdf or printed})"

    def _page_anchor(self, page: Any) -> str:
        if page in (None, ""):
            return ""
        return f"#page={page}"

    def _build_pdf_uri(self, page: Any) -> str:
        if not self.admin_guide_pdf_path or page in (None, ""):
            return ""
        path = Path(self.admin_guide_pdf_path).resolve()
        return f"{path.as_uri()}#page={page}"

    def _entry_pdf_uri(self, entry: Dict[str, Any], page: Any) -> str:
        uri = str(entry.get("pdf_uri") or "")
        if "#page=" in uri:
            return uri
        if uri and page not in (None, ""):
            return f"{uri.split('#', 1)[0]}#page={page}"
        return self._build_pdf_uri(page)

    def _normalize_index_entry(self, entry: Dict[str, Any]) -> None:
        page = entry.get("pdf_page") or entry.get("page")
        if page in (None, ""):
            return
        entry["page"] = entry.get("page") or page
        entry["pdf_page"] = entry.get("pdf_page") or page
        anchor = entry.get("pdf_anchor") or entry.get("anchor") or self._page_anchor(page)
        entry["anchor"] = anchor
        entry["pdf_anchor"] = anchor
        entry["pdf_uri"] = self._entry_pdf_uri(entry, page)

    @staticmethod
    def _should_reference_text(text: str) -> bool:
        normalized = f" {text.lower()} "
        if len(normalized.strip()) < 12:
            return False
        excluded = (
            "price", "cost", "payment", "warranty", "delivery", "bidder", "boq", "tax",
            "commercial", "legal", "eligibility", "authorization", "training", "rack",
            "ups", "cooling", "cctv", "civil", "electrical",
        )
        if any(term in normalized for term in excluded):
            return False
        technical = (
            "firewall", "policy", "vpn", "ssl vpn", "ipsec", "sd-wan", "routing", "bgp",
            "ospf", "ha", "high availability", "nat", "authentication", "radius", "ldap",
            "saml", "logging", "syslog", "ips", "antivirus", "web filter", "application control",
            "ztna", "certificate", "vdom", "interface", "zone", "dns", "dhcp",
        )
        return any(term in normalized for term in technical)

    @staticmethod
    def _build_tag_from_text(text: str) -> Dict[str, Any]:
        query = re.sub(r"\s+", " ", text).strip()
        query = query[:180]
        domains = []
        lowered = query.lower()
        mapping = {
            "VPN": ("vpn", "ipsec", "ssl vpn"),
            "Firewall Policy": ("firewall", "policy", "nat"),
            "Routing": ("routing", "bgp", "ospf"),
            "High Availability": ("ha", "high availability", "cluster"),
            "Authentication": ("authentication", "ldap", "radius", "saml"),
            "Logging": ("log", "syslog", "fortianalyzer"),
            "Security Profiles": ("ips", "antivirus", "web filter", "application control"),
            "SD-WAN": ("sd-wan", "sdwan"),
        }
        for domain, keywords in mapping.items():
            if any(keyword in lowered for keyword in keywords):
                domains.append(domain)
        return {
            "reference_needed": True,
            "lookup_query": query,
            "fortinet_domains": domains,
            "priority": "medium",
        }

    # ------------------------------------------------------------------
    # 8. Enrich a single sheet
    # ------------------------------------------------------------------
    def enrich_sheet(self, ws) -> int:
        """
        Process every row in a worksheet.

        Returns:
            Number of rows enriched.
        """
        if ws.max_row is None or ws.max_row < 2:
            return 0

        # --- Discover columns ---
        headers = [str(cell.value or "").strip() for cell in ws[1]]

        tag_col_idx = None
        for i, h in enumerate(headers):
            if h == TAG_COLUMN_NAME:
                tag_col_idx = i
                break

        text_col_idx = None
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if any(kw in h_lower for kw in ("description", "requirement", "specification", "detail", "text")):
                text_col_idx = i
                break
        if text_col_idx is None:
            # Fallback: use column index 1 if it exists
            text_col_idx = min(1, len(headers) - 1)

        # --- Append output column headers ---
        if OUTPUT_COLUMNS[0] in headers:
            start_col = headers.index(OUTPUT_COLUMNS[0]) + 1
        else:
            start_col = ws.max_column + 1
            headers.append(OUTPUT_COLUMNS[0])
        for offset, col_name in enumerate(OUTPUT_COLUMNS):
            cell = ws.cell(row=1, column=start_col + offset, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # --- Process data rows ---
        enriched_count = 0
        total_data_rows = ws.max_row - 1

        for row_idx in range(2, ws.max_row + 1):
            self.stats["total_rows"] += 1

            tag_value = None
            if tag_col_idx is not None:
                tag_cell = ws.cell(row=row_idx, column=tag_col_idx + 1)
                tag_value = tag_cell.value

            req_text = ""
            if text_col_idx is not None:
                req_cell = ws.cell(row=row_idx, column=text_col_idx + 1)
                req_text = str(req_cell.value or "")

            tag_data = self.parse_reference_tag(tag_value)
            if tag_data is None and self._should_reference_text(req_text):
                tag_data = self._build_tag_from_text(req_text)

            if tag_data is None:
                self._write_not_required(ws, row_idx, start_col)
                self.stats["not_required"] += 1
                continue

            try:
                match = self.find_best_match(tag_data, req_text)
                citation = self.generate_citation(match)
                
                if match["status"] != "Matched":
                    citation = f"[Review Required] {citation}" if citation else "[Review Required] No match found"

                cell = ws.cell(row=row_idx, column=start_col)
                self._write_citation_cell(cell, citation, match)
                cell.fill = CITATION_FILL if match["status"] == "Matched" else REVIEW_FILL
                cell.border = BORDER_THIN
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                enriched_count += 1
                if match["status"] == "Matched":
                    self.stats["matched"] += 1
                else:
                    self.stats["manual_review"] += 1

            except Exception as exc:
                logger.error("Error enriching row %d in '%s': %s", row_idx, ws.title, exc)
                self._write_error(ws, row_idx, start_col, str(exc))
                self.stats["errors"] += 1

            # Progress
            if (row_idx - 1) % 25 == 0 or row_idx == ws.max_row:
                logger.info(
                    "  Sheet '%s': processed %d / %d rows",
                    ws.title, row_idx - 1, total_data_rows,
                )

        if tag_col_idx is not None:
            ws.delete_cols(tag_col_idx + 1)
            if tag_col_idx + 1 < start_col:
                start_col -= 1
            new_col_letter = ws.cell(row=1, column=start_col).column_letter
            ws.column_dimensions[new_col_letter].width = 80
        else:
            new_col_letter = ws.cell(row=1, column=start_col).column_letter
            ws.column_dimensions[new_col_letter].width = 80

        return enriched_count

    @staticmethod
    def _write_citation_cell(cell, citation: str, match: Dict[str, Any]) -> None:
        # Check for web URL first, fallback to pdf_uri
        uri = match.get("url") or match.get("pdf_uri") or ""
        
        # If it's a web URL, don't try to append PDF pages
        is_web_url = str(uri).startswith("http")
        
        if not is_web_url:
            page = match.get("pdf_page") or match.get("page")
            if uri and "#page=" not in str(uri) and page not in (None, ""):
                uri = f"{str(uri).split('#', 1)[0]}#page={page}"
                
        if citation and uri:
            safe_uri = str(uri).replace('"', '""')
            safe_label = str(citation).replace('"', '""')
            cell.value = f'=HYPERLINK("{safe_uri}","{safe_label}")'
            cell.font = Font(color="0563C1", underline="single")
        else:
            cell.value = citation

    def _write_not_required(self, ws, row_idx: int, start_col: int) -> None:
        """Fill output columns for a row that does not need a reference."""
        cell = ws.cell(row=row_idx, column=start_col)
        cell.value = ""
        cell.fill = NOT_REQ_FILL
        cell.border = BORDER_THIN
        cell.alignment = Alignment(vertical="top")

    def _write_error(self, ws, row_idx: int, start_col: int, error_msg: str) -> None:
        """Fill output columns for a row where processing failed."""
        cell = ws.cell(row=row_idx, column=start_col)
        cell.value = f"[ERROR] {error_msg}"
        cell.fill = REVIEW_FILL
        cell.border = BORDER_THIN
        cell.alignment = Alignment(vertical="top")

    # ------------------------------------------------------------------
    # 9. Enrich entire workbook
    # ------------------------------------------------------------------
    def enrich_workbook(self) -> None:
        """Iterate all sheets and enrich rows with Admin Guide references."""
        if self.wb is None:
            raise RuntimeError("Workbook not loaded. Call load_workbook() first.")
        if not self.toc_index:
            raise RuntimeError("TOC index not loaded. Call load_toc_index() first.")

        logger.info("=== Starting workbook enrichment ===")
        start_time = time.time()

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            logger.info("Processing sheet: '%s' (%d rows)", sheet_name, ws.max_row or 0)
            enriched = self.enrich_sheet(ws)
            logger.info("  → Enriched %d rows in '%s'", enriched, sheet_name)

        elapsed = time.time() - start_time
        logger.info("=== Enrichment complete in %.1fs ===", elapsed)
        logger.info(
            "Stats: total=%d, matched=%d, review=%d, not_required=%d, errors=%d",
            self.stats["total_rows"],
            self.stats["matched"],
            self.stats["manual_review"],
            self.stats["not_required"],
            self.stats["errors"],
        )

    # ------------------------------------------------------------------
    # 10. Save output
    # ------------------------------------------------------------------
    def save_output(self, output_path: Optional[str] = None) -> str:
        """
        Save the enriched workbook.

        If no output_path is given, appends ``_enriched`` to the input filename.
        """
        if self.wb is None:
            raise RuntimeError("No workbook to save.")

        if output_path is None:
            base, ext = os.path.splitext(self.workbook_path)
            output_path = f"{base}_enriched_with_admin_guide_references{ext}"

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        self._rewrite_admin_guide_links_for_excel(output_path)
        self.wb.save(output_path)
        logger.info("Enriched workbook saved to: %s", output_path)
        return output_path

    def _rewrite_admin_guide_links_for_excel(self, output_path: str) -> None:
        if self.wb is None:
            return
        
        server_base_url = os.getenv("SERVER_BASE_URL")
        if server_base_url and server_base_url.endswith("/"):
            server_base_url = server_base_url[:-1]
            
        link_dir = os.path.join(os.path.dirname(output_path) or ".", "admin_guide_page_links")
        
        for ws in self.wb.worksheets:
            headers = [str(cell.value or "").strip() for cell in ws[1]]
            if OUTPUT_COLUMNS[0] not in headers:
                continue
            col_idx = headers.index(OUTPUT_COLUMNS[0]) + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                parsed = self._parse_hyperlink_formula(cell.value)
                if parsed is None:
                    continue
                uri, label = parsed
                if not uri.lower().startswith("file:///") or ".pdf#page=" not in uri.lower():
                    continue
                
                if server_base_url:
                    page_match = re.search(r"#page=(\d+)", uri, re.IGNORECASE)
                    page = page_match.group(1) if page_match else "1"
                    new_uri = f"{server_base_url}/reference_pdf#page={page}"
                else:
                    new_uri = self._create_pdf_jump_file(link_dir, uri)
                    
                cell.value = f'=HYPERLINK("{self._excel_quote(new_uri)}","{self._excel_quote(label)}")'

    @staticmethod
    def _parse_hyperlink_formula(value: Any) -> Optional[Tuple[str, str]]:
        if not isinstance(value, str):
            return None
        match = re.match(r'^=HYPERLINK\("((?:[^"]|"")*)","((?:[^"]|"")*)"\)$', value)
        if not match:
            return None
        return match.group(1).replace('""', '"'), match.group(2).replace('""', '"')

    @staticmethod
    def _excel_quote(value: str) -> str:
        return str(value).replace('"', '""')

    @staticmethod
    def _create_pdf_jump_file(link_dir: str, pdf_uri: str) -> str:
        os.makedirs(link_dir, exist_ok=True)
        page_match = re.search(r"#page=(\d+)", pdf_uri, re.IGNORECASE)
        page = page_match.group(1) if page_match else "1"
        target = pdf_uri
        filename = f"fortios_admin_guide_page_{page}.html"
        path = os.path.join(link_dir, filename)
        html_body = (
            "<!doctype html><html><head><meta charset=\"utf-8\">"
            f"<meta http-equiv=\"refresh\" content=\"0; url={html.escape(target, quote=True)}\">"
            f"<script>window.location.replace({json.dumps(target)});</script>"
            f"<title>FortiOS Admin Guide Page {html.escape(page)}</title></head>"
            f"<body><a href=\"{html.escape(target, quote=True)}\">Open FortiOS Admin Guide page {html.escape(page)}</a></body></html>"
        )
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(html_body)
        return Path(path).resolve().as_uri()


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------
def main() -> None:
    import argparse

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    parser = argparse.ArgumentParser(
        description="Enrich RFP Excel workbook with Fortinet Admin Guide references."
    )
    parser.add_argument("--input", required=True, help="Path to the input Excel workbook")
    parser.add_argument("--toc", default=os.path.join("output", "toc_flat_index.json"),
                        help="Path to the flat TOC index JSON")
    parser.add_argument("--output", default=None,
                        help="Output Excel path (default: <input>_enriched_with_admin_guide_references.xlsx)")
    parser.add_argument("--model", default="all-MiniLM-L6-v2",
                        help="Sentence-transformer model name")
    parser.add_argument("--pdf", default=None,
                        help="Path to the FortiOS Administration Guide PDF for file:// page hyperlinks")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: Input file not found: {args.input}")
        return
    if not os.path.exists(args.toc):
        print(f"ERROR: TOC index not found: {args.toc}")
        print("Run: python scripts/pdf_admin_metadata.py --pdf \"data/Reference dataset/FortiOS-7.6.6-Administration_Guide.pdf\"")
        return

    enricher = FortinetAdminGuideReferenceEnricher(
        workbook_path=args.input,
        toc_index_path=args.toc,
        model_name=args.model,
        admin_guide_pdf_path=args.pdf,
    )

    enricher.load_workbook()
    enricher.load_toc_index()
    enricher.build_embedding_index()
    enricher.enrich_workbook()
    output = enricher.save_output(args.output)

    print(f"\nSUCCESS: Enriched workbook saved to: {output}")
    print(f"Stats: {enricher.stats}")


if __name__ == "__main__":
    main()
